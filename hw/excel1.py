# устанавливаем библиотеку для работы с эксель
# pip install openpyxl
# импорт
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval

file_name = "samplel.xlsx"

workbook = openpyxl.load_workbook("samplel.xlsx")
worksheet = workbook.active
max_row = worksheet.max_row + 1
print(max_row)
print(type(max_row))
max_column = worksheet.max_column + 1
print(max_column)
print(type(max_column))
index = 0

workbook.save('homework1.xlsx')


