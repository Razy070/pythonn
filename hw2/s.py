# import openpyxl
#
# book = openpyxl.load_workbook("test1.xlsx")
# sheet = book.active
#
# global_rows = []
# for row in range(1, sheet.max_row + 1):
#     local_row = []
#     for column in range(1, sheet.max_column + 1):
#         local_row.append(sheet.cell(row=row, column=column).value)
#     global_rows.append(local_row)
#
# for i in global_rows:
#     print(i)

import openpyxl

workbook = openpyxl.load_workbook("test1.xlsx")
worksheet = workbook.active
max_row = worksheet.max_row + 1
max_column = worksheet.max_column + 1

external_array = []
for row in range(1, max_row + 1):
    internal_array = []
    for col in range(1, max_column + 1):
        value = worksheet.cell(row=row, column=col).value
        if value is None:
            value = ""
        internal_array.append(value)
    external_array.append(internal_array)
print(external_array)

set1 = set(external_array[1])
set1.remove('')
print(set1)
set2 = set(external_array[3])
set2.remove('')
print(set2)
set3 = set(external_array[5])
set3.remove('')
print(set3)

set4 = set2.intersection(set1, set3)
print(set4)

set5 = set2.difference(set1, set3)
print(set5)

set6 = set3.union(set1)
print(set6)

