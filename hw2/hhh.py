# устанавливаем библиотеку для работы с эксель
# pip install openpyxl
# импорт
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval

file_name = "test1.xlsx"

workbook = openpyxl.load_workbook(file_name)
worksheet = workbook.active

rows = worksheet.max_row
cols = worksheet.max_column


for i in range(1, rows + 1):
    string = ''
    for j in range(1, cols + 1):
        cell = worksheet.cell(row = i, column = j)
        string = string + str(cell.value) + ' '
    print(string)












