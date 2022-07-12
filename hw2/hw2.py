import openpyxl

book = openpyxl.load_workbook("test1.xlsx")
sheet = book.active

global_rows = []
for row in range(1, sheet.max_row + 1):
    local_row = []
    for column in range(1, sheet.max_column + 1):
        local_row.append(sheet.cell(row=row, column=column).value)
    global_rows.append(local_row)

for i in global_rows:
    print(i)






