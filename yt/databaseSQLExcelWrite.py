import psycopg2
import openpyxl
from openpyxl.utils import get_column_letter

# Read all data from SQL
conn = psycopg2.connect("dbname=user=postgres password=R29062011Z")

cur = conn.cursor()
cur.execute("""
SELECT * FROM public.table1
""")
records = cur.fetchall()
conn.close()

# Write all data to .xlsx
book = openpyxl.Workbook()
sheet = book.active

titles = ["Заголовок 1", "Заголовок 2", "Заголовок 3"]
index_col = 1
for title in titles:
    sheet[f"{get_column_letter(index_col)}1"] = str(title)
    index_col += 1

row_index = 2
for row in records:
    col_index = 1
    for column in row:
        sheet[f"{get_column_letter(col_index)}{row_index}"] = str(column)
        col_index += 1
    row_index += 1
book.save('excel_data/new_table1.xlsx')
