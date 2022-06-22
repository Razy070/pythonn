
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from home__work4 import j

kniga1 = Workbook()
stranica1 = kniga1.active

var_range = range(1, 20)
stroka_1 = []
stroka1 = "A_12"
for i in var_range:
    stroka1 = f"{i}"
    stroka_1.append(stroka1)

stroka_2 = []
stroka2 = ""
for i in var_range:
    stroka2 = f"A_{j}"
    stroka_2.append(stroka2)

stroka_finall = []
stroka_finall.append(stroka2)
stroka_finall.append(stroka2)

for row in range(0, len(stroka_finall)):
    for col in range(0, len(stroka_finall[row])):


        stranica1[f"{get_column_letter(col + 1)}{row + 1}"]=stroka_finall[row][col]

for col in range(0, len(stroka_finall)):
    for row in range(0 , len(stroka_finall[col])):
        print(row)
        if get_column_letter(col + 1) == "A":
            letter = "B"
        else:
            letter = "A"
        stranica1[f"{letter}{row + 4}"] = stroka_finall[col][row]
print(stroka_finall)


kniga1.save("temp/homework44.xlsx")
